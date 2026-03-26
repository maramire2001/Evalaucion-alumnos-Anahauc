from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse, JsonResponse
from django.conf import settings
from .forms import SubmissionForm
from .models import Student, Submission, Grade
from .grading import grade_submission
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import zipfile
import io
import os
import json


# ──────────────────────────────────────────────
#  VISTA ESTUDIANTE: Entregar trabajo
# ──────────────────────────────────────────────
def submit_assignment(request):
    if request.method == 'POST':
        form = SubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            name = form.cleaned_data['name']
            student_id = form.cleaned_data['student_id']
            blog_url = form.cleaned_data['blog_url']
            pdf_file = form.cleaned_data['pdf_file']

            student, created = Student.objects.get_or_create(
                student_id=student_id,
                defaults={'name': name}
            )
            if not created and student.name != name:
                student.name = name
                student.save()

            # Leer PDF en memoria y guardar en DB (PostgreSQL permanente)
            pdf_bytes = pdf_file.read() if pdf_file else None
            pdf_name = pdf_file.name if pdf_file else ''
            submission = Submission.objects.create(
                student=student,
                blog_url=blog_url,
                pdf_data=pdf_bytes,
                pdf_filename=pdf_name,
            )

            # Grade automatically
            grade = grade_submission(submission)
            grade.save()

            messages.success(request, 'Entrega enviada y calificada exitosamente.')
            return redirect('submit')
    else:
        form = SubmissionForm()
    return render(request, 'submit.html', {'form': form})


# ──────────────────────────────────────────────
#  PANEL PROFESOR
# ──────────────────────────────────────────────
@login_required(login_url='/admin/login/')
def profesor_dashboard(request):
    """Panel principal del profesor con todas las entregas y calificaciones."""
    grades = Grade.objects.select_related(
        'submission__student'
    ).order_by('-final_score')

    # Parsear JSON completo (incluye scores por criterio y justificaciones)
    grade_data = []
    for g in grades:
        try:
            datos = json.loads(g.adjustments) if g.adjustments else {}
        except (json.JSONDecodeError, TypeError):
            datos = {}
        grade_data.append({
            'grade': g,
            'justificaciones': datos,   # contiene scores, entradas_detectadas y justificaciones
        })

    # Estadísticas resumen
    total_alumnos = Student.objects.count()
    total_entregas = Submission.objects.count()
    promedio = sum(g.grade.final_score for g in grade_data) / len(grade_data) if grade_data else 0

    context = {
        'grade_data': grade_data,
        'total_alumnos': total_alumnos,
        'total_entregas': total_entregas,
        'promedio': round(promedio, 2),
    }
    return render(request, 'profesor_dashboard.html', context)


@login_required(login_url='/admin/login/')
def download_pdf(request, submission_id):
    """Descarga el PDF de una entrega específica."""
    submission = get_object_or_404(Submission, id=submission_id)
    # Primero intentar desde la DB (nuevo sistema permanente)
    if submission.pdf_data:
        student_name = submission.student.name.replace(' ', '_')
        filename = submission.pdf_filename or f"trabajo_{student_name}.pdf"
        response = HttpResponse(bytes(submission.pdf_data), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    # Fallback: filesystem (solo funciona si no hubo reinicio)
    if submission.pdf_file:
        try:
            file_path = submission.pdf_file.path
            if os.path.exists(file_path):
                student_name = submission.student.name.replace(' ', '_')
                filename = f"trabajo_{student_name}.pdf"
                response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        except Exception:
            pass
    return HttpResponse("PDF no encontrado. El alumno debe volver a subir su entrega.", status=404)


@login_required(login_url='/admin/login/')
def download_all_pdfs(request):
    """Descarga todos los PDFs en un archivo ZIP."""
    submissions = Submission.objects.select_related('student').all()
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sub in submissions:
            student_name = sub.student.name.replace(' ', '_')
            arcname = f"{student_name}_{sub.student.student_id}.pdf"
            if sub.pdf_data:
                zf.writestr(arcname, bytes(sub.pdf_data))
            elif sub.pdf_file:
                try:
                    fp = sub.pdf_file.path
                    if os.path.exists(fp):
                        zf.write(fp, arcname)
                except Exception:
                    pass
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="todos_los_trabajos.zip"'
    return response


@login_required(login_url='/admin/login/')
def pdf_portada(request, submission_id):
    """Genera y devuelve una imagen de la portada (primera página) del PDF."""
    submission = get_object_or_404(Submission, id=submission_id)
    file_path = submission.pdf_file.path
    if not os.path.exists(file_path):
        return HttpResponse("PDF no encontrado", status=404)

    try:
        from pdf2image import convert_from_path
        import io as _io
        pages = convert_from_path(file_path, dpi=120, first_page=1, last_page=1)
        if pages:
            img_buffer = _io.BytesIO()
            pages[0].save(img_buffer, format='JPEG', quality=85)
            img_buffer.seek(0)
            return HttpResponse(img_buffer.read(), content_type='image/jpeg')
    except Exception as e:
        pass
    return HttpResponse("No se pudo generar la vista previa", status=500)


@login_required(login_url='/admin/login/')
def recalcular_calificacion(request, submission_id):
    """Re-ejecuta el scraping y calificación IA para una entrega."""
    if request.method != 'POST':
        return HttpResponse("Método no permitido", status=405)
    submission = get_object_or_404(Submission, id=submission_id)
    try:
        # Eliminar calificación previa si existe
        Grade.objects.filter(submission=submission).delete()
        # Recalcular
        grade = grade_submission(submission)
        grade.save()
        return JsonResponse({'ok': True, 'final_score': round(grade.final_score, 2)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# ──────────────────────────────────────────────
#  EXPORTAR EXCEL (mejorado)
# ──────────────────────────────────────────────
@login_required(login_url='/admin/login/')
def export_grades(request):
    wb = openpyxl.Workbook()

    # ── Hoja 1: Calificaciones globales ──
    ws = wb.active
    ws.title = "Calificaciones Globales"

    # Estilos
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['#', 'Alumno', 'ID', 'Cant./Regularidad (35%)',
               'Calidad (35%)', 'Presentación (10%)', 'Protocolo (20%)',
               'Calificación Final', 'Blog URL', 'Fecha Entrega']
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Ajuste de anchos
    col_widths = [5, 30, 15, 22, 15, 17, 17, 18, 45, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    grades = Grade.objects.select_related('submission__student').order_by('-final_score')
    for idx, g in enumerate(grades, 1):
        try:
            j = json.loads(g.adjustments) if g.adjustments else {}
        except Exception:
            j = {}
        # Reconstruir scores individuales desde el JSON
        cant = j.get('cantidad_score', '')
        cal = j.get('calidad_score', '')
        pres = j.get('presentacion_score', '')
        prot = j.get('protocolo_score', '')
        # Si no están en el nivel raíz, buscar en el campo que guardamos
        # (el grading.py guarda 'justificaciones', no los scores directamente en adjustments)
        # Los scores reales son: blog_score ≈ promedio de 3, entries_score = protocolo, total = final
        row = [
            idx,
            g.submission.student.name,
            g.submission.student.student_id,
            cant if cant != '' else round(g.blog_score, 2),
            cal if cal != '' else '',
            pres if pres != '' else '',
            prot if prot != '' else round(g.entries_score, 2),
            round(g.final_score, 2),
            g.submission.blog_url,
            g.submission.submitted_at.strftime('%d/%m/%Y %H:%M') if g.submission.submitted_at else '',
        ]
        ws.append(row)
        row_fill = alt_fill if idx % 2 == 0 else None
        for col, cell in enumerate(ws[idx + 1], 1):
            cell.border = thin
            cell.alignment = center
            if col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if row_fill:
                cell.fill = row_fill
            # Resaltar calificación final
            if col == 8:
                val = cell.value
                if isinstance(val, (int, float)):
                    if val >= 8:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(bold=True, color="166534")
                    elif val >= 6:
                        cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                        cell.font = Font(bold=True, color="854D0E")
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(bold=True, color="991B1B")

    # Fila de promedio
    ws.append([])
    prom_row_idx = ws.max_row + 1
    avg = Grade.objects.aggregate(avg=__import__('django.db.models', fromlist=['Avg']).Avg('final_score'))['avg']
    ws.cell(prom_row_idx, 2, "PROMEDIO GRUPO").font = Font(bold=True)
    ws.cell(prom_row_idx, 8, round(avg, 2) if avg else 0).font = Font(bold=True)
    ws.cell(prom_row_idx, 8).fill = orange_fill
    ws.cell(prom_row_idx, 8).font = Font(bold=True, color="FFFFFF")

    # ── Hoja 2: Justificaciones IA ──
    ws2 = wb.create_sheet("Retroalimentación IA")
    ws2.append(['Alumno', 'ID', 'Calific. Final',
                'Justif. Cantidad', 'Justif. Calidad',
                'Justif. Presentación', 'Justif. Protocolo'])
    for col, cell in enumerate(ws2[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    for w, col in zip([30, 15, 14, 50, 50, 50, 50], 'ABCDEFG'):
        ws2.column_dimensions[col].width = w

    for g in grades:
        try:
            j = json.loads(g.adjustments) if g.adjustments else {}
        except Exception:
            j = {}
        just = j.get('justificaciones', {})
        ws2.append([
            g.submission.student.name,
            g.submission.student.student_id,
            round(g.final_score, 2),
            just.get('cantidad', ''),
            just.get('calidad', ''),
            just.get('presentacion', ''),
            just.get('protocolo', ''),
        ])
        for cell in ws2[ws2.max_row]:
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=calificaciones_maestria.xlsx'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
#  REPORTE DE SUBIDAS
# ──────────────────────────────────────────────
@login_required(login_url='/admin/login/')
def reporte_subidas(request):
    """Vista con tabla de quién subió, cuándo y qué."""
    submissions = Submission.objects.select_related('student').order_by('-submitted_at')
    return render(request, 'reporte_subidas.html', {'submissions': submissions})


@login_required(login_url='/admin/login/')
def export_subidas_excel(request):
    """Exporta el reporte de subidas a Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Subidas"

    # Header
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:G1')
    ws['A1'] = 'REPORTE DE SUBIDAS — Seminario de Integración Anáhuac 2026'
    ws['A1'].font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    headers = ['#', 'Alumno', 'ID del Alumno', 'Fecha de Subida', 'Hora',
               'URL del Blog', 'Archivo PDF']
    ws.append(headers)
    for col, cell in enumerate(ws[2], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    col_widths = [5, 30, 15, 16, 10, 50, 35]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    submissions = Submission.objects.select_related('student').order_by('-submitted_at')
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    for idx, sub in enumerate(submissions, 1):
        fecha = sub.submitted_at.strftime('%d/%m/%Y') if sub.submitted_at else ''
        hora  = sub.submitted_at.strftime('%H:%M:%S') if sub.submitted_at else ''
        pdf_nombre = str(sub.pdf_file).split('/')[-1] if sub.pdf_file else '—'
        row = [idx, sub.student.name, sub.student.student_id,
               fecha, hora, sub.blog_url, pdf_nombre]
        ws.append(row)
        r = ws.max_row
        ws.row_dimensions[r].height = 22
        for col_i, cell in enumerate(ws[r], 1):
            cell.border = thin
            cell.alignment = Alignment(
                horizontal='left' if col_i in [2, 6, 7] else 'center',
                vertical='center'
            )
            cell.font = Font(size=10, name="Arial")
            if idx % 2 == 0:
                cell.fill = alt_fill

    # Fila de totales
    ws.append([])
    tr = ws.max_row + 1
    ws.cell(tr, 2, f"TOTAL: {submissions.count()} entregas").font = Font(bold=True, name="Arial")
    ws.cell(tr, 2).fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    ws.cell(tr, 2).font = Font(bold=True, color="FFFFFF", name="Arial")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_subidas.xlsx"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
#  SETUP: Crear superusuario (URL secreta de un solo uso)
# ──────────────────────────────────────────────
def setup_admin(request):
    """
    URL secreta para crear/restablecer el superusuario del profesor.
    Visita: /init-profesor-anahuac2026/
    """
    from django.contrib.auth.models import User
    USERNAME = 'profesor'
    PASSWORD = 'AnAhuac2026!'
    EMAIL = 'marioantonioramirezbarajas@gmail.com'

    try:
        if User.objects.filter(username=USERNAME).exists():
            u = User.objects.get(username=USERNAME)
            u.set_password(PASSWORD)
            u.is_superuser = True
            u.is_staff = True
            u.save()
            msg = f"✅ Contraseña del usuario '{USERNAME}' restablecida correctamente."
        else:
            User.objects.create_superuser(USERNAME, EMAIL, PASSWORD)
            msg = f"✅ Superusuario '{USERNAME}' creado correctamente."

        html = f"""
        <html><body style="font-family:Arial;text-align:center;padding:60px;background:#0f172a;color:#e2e8f0;">
        <h1 style="color:#4ade80;">{msg}</h1>
        <p style="font-size:1.2em;">Usuario: <b>{USERNAME}</b><br>Contraseña: <b>{PASSWORD}</b></p>
        <p><a href="/admin/" style="color:#60a5fa;font-size:1.1em;">→ Ir al Admin</a> &nbsp;|&nbsp;
           <a href="/profesor/" style="color:#60a5fa;font-size:1.1em;">→ Panel del Profesor</a></p>
        <p style="color:#64748b;margin-top:40px;font-size:0.85em;">
        Por seguridad, puedes eliminar la URL <code>/init-profesor-anahuac2026/</code> de urls.py cuando ya no la necesites.
        </p>
        </body></html>
        """
        return HttpResponse(html)
    except Exception as e:
        return HttpResponse(f"<html><body style='color:red;padding:40px;'><h2>Error: {e}</h2></body></html>",
                            status=500)
